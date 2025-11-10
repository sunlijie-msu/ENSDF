import openpyxl
import sys

wb = openpyxl.load_workbook(r'd:\X\ND\ENSDF\Statistics.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    print(f'\n=== Sheet: {sheet_name} ===')
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        print(row)
